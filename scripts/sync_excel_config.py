import openpyxl
import os

def sync_excel_data(file_path):
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return

    print(f"Loading {file_path}...")
    # data_only=False にして、数式ではなく値を保持したまま書き込み可能にする
    wb = openpyxl.load_workbook(file_path)
    
    if 'プレミアムプラン用' not in wb.sheetnames or 'チェックリスト' not in wb.sheetnames:
        print("Error: Required sheets not found.")
        return

    src_sheet = wb['プレミアムプラン用']
    dst_sheet = wb['チェックリスト']

    keywords = [
        "医院名",
        "URL",
        "住所",
        "院長名・副院長名",
        "電話番号",
        "診療時間",
        "敬称統一表記",
        "GA4コード"
    ]

    # データ抽出用の辞書
    data = {}

    def find_value_to_right(sheet, keyword):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).replace('\n', '').strip() == keyword:
                    # キーワードのすぐ右のセルを取得
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    return target_cell.value
        return None

    print("Extracting data from 'プレミアムプラン用'...")
    for k in keywords:
        val = find_value_to_right(src_sheet, k)
        data[k] = val
        print(f"  {k}: {val}")

    def set_value_to_right(sheet, keyword, value):
        if value is None:
            return False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).replace('\n', '').strip() == keyword:
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    target_cell.value = value
                    return True
        return False

    print("Writing data to 'チェックリスト'...")
    for k, v in data.items():
        if set_value_to_right(dst_sheet, k, v):
            print(f"  Updated {k}")
        else:
            print(f"  Warning: Keyword '{k}' not found in 'チェックリスト'")

    print(f"Saving {file_path}...")
    wb.save(file_path)
    print("Sync completed successfully.")

if __name__ == "__main__":
    sync_excel_data('DC-config.xlsx')
