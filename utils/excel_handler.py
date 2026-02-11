import openpyxl
import os

class ExcelHandler:
    """DC-config.xlsxの操作を担当するクラス"""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = None
        
    def load(self):
        """ファイルをロード"""
        if os.path.exists(self.file_path):
            self.wb = openpyxl.load_workbook(self.file_path)
            return True
        return False

    def sync_sheets(self):
        """
        'プレミアムプラン用' から 'チェックリスト' へデータを同期
        改行を無視してマッチングを行う
        """
        if not self.wb or 'プレミアムプラン用' not in self.wb.sheetnames or 'チェックリスト' not in self.wb.sheetnames:
            return False

        src_sheet = self.wb['プレミアムプラン用']
        dst_sheet = self.wb['チェックリスト']

        keywords = [
            "医院名", "URL", "住所", "院長名・副院長名",
            "電話番号", "診療時間", "敬称統一表記", "GA4コード"
        ]

        data = {}
        # 抽出
        for kw in keywords:
            data[kw] = self._find_value_to_right(src_sheet, kw)

        # 反映
        for kw, val in data.items():
            if val is not None:
                self._set_value_to_right(dst_sheet, kw, val)

        self.wb.save(self.file_path)
        return True

    def get_basic_info(self):
        """
        'チェックリスト' シートから基本3点(URL, 医院名, 電話番号)を取得
        """
        if not self.wb or 'チェックリスト' not in self.wb.sheetnames:
            return None, None, None

        sheet = self.wb['チェックリスト']
        url = self._find_value_to_right(sheet, "URL")
        clinic_name = self._find_value_to_right(sheet, "医院名")
        phone = self._find_value_to_right(sheet, "電話番号")

        return url, clinic_name, phone

    def _find_value_to_right(self, sheet, keyword):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).replace('\n', '').strip() == keyword:
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    return target_cell.value
        return None

    def _set_value_to_right(self, sheet, keyword, value):
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).replace('\n', '').strip() == keyword:
                    target_cell = sheet.cell(row=cell.row, column=cell.column + 1)
                    target_cell.value = value
                    return True
        return False
