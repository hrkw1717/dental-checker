"""
Excelレポート生成

チェック結果をExcelファイルに出力
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict
from io import BytesIO


class ExcelReporter:
    """Excelレポート生成クラス"""
    
    def __init__(self, config: Dict):
        """
        Args:
            config: 設定辞書
        """
        self.config = config
        self.output_config = config.get("output", {}).get("excel", {})
        self.columns = self.output_config.get("columns", [
            "No", "ページ", "チェック項目", "結果", "詳細", "重要度"
        ])
        self.result_symbols = self.output_config.get("result_symbols", {
            "ok": "✅",
            "warning": "⚠️",
            "error": "❌"
        })
    
    def generate_report(self, clinic_name: str, results: List[Dict]) -> BytesIO:
        """
        チェック結果からExcelレポートを生成
        
        Args:
            clinic_name: クリニック名
            results: チェック結果のリスト
        
        Returns:
            ExcelファイルのBytesIO
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "チェック結果"
        
        # ヘッダー行を作成
        self._create_header(ws)
        
        # データ行を追加
        for idx, result in enumerate(results, start=1):
            self._add_result_row(ws, idx, result)
        
        # 列幅を調整
        self._adjust_column_widths(ws)
        
        # BytesIOに保存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_header(self, ws):
        """ヘッダー行を作成"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, column_name in enumerate(self.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _add_result_row(self, ws, row_num: int, result: Dict):
        """結果行を追加"""
        row_idx = row_num + 1  # ヘッダー行の次から
        
        # 結果ステータスに応じた記号
        status_symbol = self.result_symbols.get(result["status"], result["status"])
        
        # 行データ
        row_data = [
            row_num,
            result.get("page_url", ""),
            result.get("check_name", ""),
            status_symbol,
            result.get("details", ""),
            result.get("severity", "")
        ]
        
        # セルに値を設定
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # ステータスに応じて背景色を設定
            if col_idx == 4:  # 結果列
                if result["status"] == "ok":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif result["status"] == "warning":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif result["status"] == "error":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _adjust_column_widths(self, ws):
        """列幅を調整"""
        column_widths = {
            "No": 8,
            "ページ": 40,
            "チェック項目": 20,
            "結果": 10,
            "詳細": 60,
            "重要度": 12
        }
        
        for col_idx, column_name in enumerate(self.columns, start=1):
            width = column_widths.get(column_name, 15)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
